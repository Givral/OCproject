import cv2
import numpy as np
from math import atan
import pandas
from openpyxl import load_workbook
import array

import os
os.chdir('E:/pictures/bash')
book = load_workbook('datatop0.xlsx')

# capture video
cap = cv2.VideoCapture("top0.mp4")
cap1 = cv2.VideoCapture("top0pano.mp4")
cap2 = cv2.VideoCapture("top0cut.mp4")
cap3 = cv2.VideoCapture("top0radius70.mp4")

#load xlsx
#book = load_workbook('datatop0.xlsx')

# function to create BGS
subtractorKNN = cv2.createBackgroundSubtractorKNN()
subtractorKNN1 = cv2.createBackgroundSubtractorKNN()
subtractorKNN2 = cv2.createBackgroundSubtractorKNN()
subtractorKNN3 = cv2.createBackgroundSubtractorKNN()

# function to create morphology
kernel1 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
kernel2 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
# write header
h_ratio_w_h ='ratio_w_h'
h_area = 'area'
h_area_rect = 'area_rect'
h_ratio_area = 'ratio_area'
h_angle_of_rect = 'angle_of_rect'
h_angle_o_line = 'angle_o_line'
h_camera_region = 'camera_region'

#writing file header
header = pandas.DataFrame({'ratio_w_h':[h_ratio_w_h], 'area':[h_area], 'area_rect':[h_area_rect], 'ratio_area':[h_ratio_area],
            'angle_of_rect':[h_angle_of_rect], 'angle_o_line':[h_angle_o_line],'camera_region':[h_camera_region]})
writer = pandas.ExcelWriter('datatop0.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
header.to_excel(writer, startrow=1, header=False, index=False)
writer.save()

while True:
    for i in range(2, 1000, 1):
        _, frame = cap.read()
        _, frame1 = cap1.read()
        _, frame2 = cap2.read()
        _, frame3 = cap3.read()

        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gray_frame1 = cv2.cvtColor(frame1, cv2.COLOR_BGR2GRAY)
        gray_frame2 = cv2.cvtColor(frame2, cv2.COLOR_BGR2GRAY)
        gray_frame3 = cv2.cvtColor(frame3, cv2.COLOR_BGR2GRAY)

        gaussian_blur = cv2.GaussianBlur(gray_frame, (3, 3), 0)
        gaussian_blur1 = cv2.GaussianBlur(gray_frame1, (3, 3), 0)
        gaussian_blur2 = cv2.GaussianBlur(gray_frame2, (3, 3), 0)
        gaussian_blur3 = cv2.GaussianBlur(gray_frame3, (3, 3), 0)

        mask = subtractorKNN.apply(gaussian_blur)
        mask1 = subtractorKNN1.apply(gaussian_blur1)
        mask2 = subtractorKNN2.apply(gaussian_blur2)
        mask3 = subtractorKNN3.apply(gaussian_blur3)

        _, threshold = cv2.threshold(mask, 200, 255, cv2.THRESH_BINARY)
        _, threshold1 = cv2.threshold(mask1, 200, 255, cv2.THRESH_BINARY)
        _, threshold2 = cv2.threshold(mask2, 200, 255, cv2.THRESH_BINARY)
        _, threshold3 = cv2.threshold(mask3, 200, 255, cv2.THRESH_BINARY)

        opening_m = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, kernel1, iterations=1)
        closing_m = cv2.morphologyEx(opening_m, cv2.MORPH_CLOSE, kernel2, iterations=3)

        opening_m1 = cv2.morphologyEx(threshold1, cv2.MORPH_OPEN, kernel1, iterations=1)
        closing_m1 = cv2.morphologyEx(opening_m1, cv2.MORPH_CLOSE, kernel2, iterations=3)

        opening_m2 = cv2.morphologyEx(threshold2, cv2.MORPH_OPEN, kernel1, iterations=1)
        closing_m2 = cv2.morphologyEx(opening_m2, cv2.MORPH_CLOSE, kernel2, iterations=3)

        opening_m3 = cv2.morphologyEx(threshold3, cv2.MORPH_OPEN, kernel1, iterations=1)
        closing_m3 = cv2.morphologyEx(opening_m3, cv2.MORPH_CLOSE, kernel2, iterations=3)

        contours, _ = cv2.findContours(closing_m, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
        contours1, _ = cv2.findContours(closing_m1, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
        contours2, _ = cv2.findContours(closing_m2, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
        contours3, _ = cv2.findContours(closing_m3, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

        for c1 in contours1:
            check = 0
            area1 = cv2.contourArea(c1)
            for c3 in contours3:
                area3 = cv2.contourArea(c3)
                if area3 > 30:
                    check = 1
            if check == 1:
                break
            else:
                if area1 > 150:
                    cv2.drawContours(closing_m1, c1, -1, (255, 255, 255), 2)
                    # draw rotated rect and calculate area of rotated rect
                    rect = cv2.minAreaRect(c1)
                    box = cv2.boxPoints(rect)
                    box = np.int0(box)
                    cv2.drawContours(closing_m1, [box], 0, (255, 255, 255), 1)
                    (x, y), (width, height), rect_angle = rect
                    ratio = height / width
                    area_rect = width * height
                    ratio_area = area / area_rect
                    print("ratio of w & h:", ratio)
                    print("area:", area)
                    print("area_rect", area_rect)
                    print("ratio_area:", ratio_area)
                    angle = 90 - rect_angle if (width < height) else -rect_angle
                    print("angle of rect:", angle)
                    rows, cols = closing_m1.shape[:2]
                    [vx, vy, x, y] = cv2.fitLine(c1, cv2.DIST_L2, 0, 0.01, 0.01)
                    lefty = int((-x * vy / vx) + y)
                    righty = int(((cols - x) * vy / vx) + y)
                    cv2.line(closing_m1, (cols - 1, righty), (0, lefty), (255, 255, 255))
                    angle_line = atan((abs(righty - lefty)) / (abs(cols - 1)))
                    print("angle of line", angle_line)
                    camera_type = 'pano'
                    data = pandas.DataFrame(
                        {'ratio_w_h': [ratio], 'area': [area], 'area_rect': [area_rect], 'ratio_area': [ratio_area],
                         'angle_of_rect': [angle], 'angle_o_line': [angle_line],'camera_type':[camera_type]})
                    writer = pandas.ExcelWriter('datatop0.xlsx', engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    data.to_excel(writer, startrow=i, header=False, index=False)
                    writer.save()

        for c3 in contours3:
                area3 = cv2.contourArea(c3)
                if area3 > 30:
                    for c in contours:
                        area = cv2.contourArea(c)
                        if area > 300:
                            cv2.drawContours(closing_m, c, -1, (255, 255, 255), 2)
                            # draw rotated rect and calculate area of rotated rect
                            rect = cv2.minAreaRect(c)
                            box = cv2.boxPoints(rect)
                            box = np.int0(box)
                            cv2.drawContours(closing_m, [box], 0, (255, 255, 255), 1)
                            (x, y), (width, height), rect_angle = rect
                            ratio = height / width
                            area_rect = width * height
                            ratio_area = area / area_rect
                            print("ratio of w & h:", ratio)
                            print("area:", area)
                            print("area_rect", area_rect)
                            print("ratio_area:", ratio_area)
                            angle = 90 - rect_angle if (width < height) else -rect_angle
                            print("angle of rect:", angle)
                            rows, cols = closing_m.shape[:2]
                            [vx, vy, x, y] = cv2.fitLine(c, cv2.DIST_L2, 0, 0.01, 0.01)
                            lefty = int((-x * vy / vx) + y)
                            righty = int(((cols - x) * vy / vx) + y)
                            cv2.line(closing_m, (cols - 1, righty), (0, lefty), (255, 255, 255))
                            angle_line = atan((abs(righty - lefty)) / (abs(cols - 1)))
                            print("angle of line", angle_line)
                            camera_type1 = 'fisheye'
                            data = pandas.DataFrame(
                                {'ratio_w_h': [ratio], 'area': [area], 'area_rect': [area_rect],
                                 'ratio_area': [ratio_area],
                                 'angle_of_rect': [angle], 'angle_o_line': [angle_line],'camera_type':[camera_type1]})
                            writer = pandas.ExcelWriter('datatop0.xlsx', engine='openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            data.to_excel(writer, startrow=i, header=False, index=False)
                            writer.save()


        cv2.imshow("top0",closing_m)
        cv2.imshow("top0pano", closing_m1)
        cv2.imshow("top0radius70", closing_m3)
        cv2.imshow("frame", frame)
        key = cv2.waitKey(30)
        if key == 27:
            break

    cap.release()
    cv2.waitKey(0) & 0xFF
    cv2.destroyAllWindows()









